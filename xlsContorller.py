import openpyxl
from datetime import datetime
from preflist import PREF_CODE

import pytz
tokyo_tz = pytz.timezone('Asia/Tokyo')

def editCollectionCondition(conditions):
    condition = f'【土地/建物】{conditions[0]}'
    condition += f'／【都道府県】{list(PREF_CODE)[int(conditions[1])-1]}'
    condition += f'／【市町村名】{conditions[2]}'
    condition += f'／【地番・家屋番号】{conditions[3]}～{conditions[4]}'
    condition += f'／【請求種別】{conditions[5]}'
    return condition

class xlsContorller():
    TEMPLETE_FILE_PATH = './output/template.xlsx'
    TEMPLETE_SHEET_NAME = 'テンプレート'
    OUTPUT_FILE_DIR_PATH = './output/'
    wb = ''
    ws = ''
    row_count = 6
    data_count = 1
    output_file_path = ''

    def __init__(self):
        # テンプレートxlsxを読込
        self.wb = openpyxl.load_workbook(self.TEMPLETE_FILE_PATH)

        # シートを指定
        self.ws = self.wb[self.TEMPLETE_SHEET_NAME]

        # コピーしたシートの名前を変更
        self.ws.title = '不動産請求情報一覧'
        
        # 処理開始時刻
        now = datetime.now(tokyo_tz)
        self.ws.cell(row=2, column=5).value = now.strftime("%Y-%m-%d_%H:%M:%S")

        # 保存先ファイル名
        self.output_file_path = f'{self.OUTPUT_FILE_DIR_PATH}output_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'

    def writeCondition(self, user_id, conditions):
        # 収集条件
        condition = f"【実行ユーザ】{user_id}／"
        condition += editCollectionCondition(conditions)
        self.ws.cell(row=1, column=5).value = condition
    
    def write(self, data):
        # データを書き込む
        for col in range(1, len(data)+1):
            self.ws.cell(row=self.row_count, column=col+1).value = data[col-1]
        self.row_count += 1

    def save(self):
        # 処理終了時刻
        self.ws.cell(row=3, column=5).value = datetime.now(tokyo_tz).strftime("%Y-%m-%d_%H:%M:%S")
        # 印刷範囲設定
        self.ws.print_area = f'A1:J{(self.row_count-1)}'
        # Excelファイルを保存
        self.wb.save(self.output_file_path)
    
    def save_with_exception(self):
        # エラー発生を記録
        self.ws.cell(row=3, column=6).value = '★エラー発生し処理中断しました。エラー対処後に再実行してください。★'
        # Excelファイルを保存
        self.save()

    def writeFudousan(self, seikyuJiko, shozaiType, shozai):
        # 請求種別／種別／所在及び地番・家屋番号
        self.write([self.data_count, seikyuJiko, shozaiType, shozai])
        self.data_count += 1

    def writeZumen(self, seikyuJiko, shozaiType, shozai, toukiDate, jiken_no, shozai_before, shozai_after):
        # 請求種別／種別／所在及び地番・家屋番号／登録年月日／事件ID／所在及び地番・家屋番号（事件前物件）／所在及び地番・家屋番号（事件後物件）
        self.write([self.data_count, seikyuJiko, shozaiType ,shozai, toukiDate, jiken_no, shozai_before, shozai_after])
        self.data_count += 1

    def writeZemenNasi(self, seikyuJiko, shozaiType, shozai, reasonZumenNashi):
        # 請求種別／種別／所在及び地番・家屋番号／図面なし理由
        self.write([self.data_count, seikyuJiko, shozaiType, shozai, '',  '',  '',  '', reasonZumenNashi])
        self.data_count += 1