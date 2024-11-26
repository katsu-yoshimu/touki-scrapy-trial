import openpyxl
# import gspread
from datetime import datetime
from preflist import PREF_CODE

import streamlit as st
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from oauth2client.service_account import ServiceAccountCredentials
import gspread

def editCollectionCondition(conditions):
    condition = f'【土地/建物】{conditions[0]}'
    condition += f'／【都道府県】{list(PREF_CODE)[int(conditions[1])-1]}'
    condition += f'／【市町村名】{conditions[2]}'
    condition += f'／【地番・家屋番号】{conditions[3]}～{conditions[4]}'
    condition += f'／【請求種別】{conditions[5]}'
    return condition

class gsContorller():
    TEMPLETE_FILE_PATH = '.\\output\\template.xlsx'
    TEMPLETE_SHEET_NAME = 'テンプレート'
    OUTPUT_FILE_DIR_PATH = '.\\output\\'
    wb = ''
    ws = ''
    row_count = 6
    data_count = 1
    output_file_path = ''
    _gsheet = None
    _ws = None

    write_data = {
            "starttime" : "",
            "conditions" : "",
            "rows" : [],
            "endtime" : "",
            "state" : ""
        }

    def __init__(self):
        # 初期値
        self.write_data = {
            "starttime" : "",
            "conditions" : "",
            "rows" : [],
            "endtime" : "",
            "state" : ""
        }
        
        # テンプレートxlsxを読込
        # self.wb = openpyxl.load_workbook(self.TEMPLETE_FILE_PATH)

        # シートを指定
        # self.ws = self.wb[self.TEMPLETE_SHEET_NAME]

        # コピーしたシートの名前を変更
        # self.ws.title = '不動産請求情報一覧'
        
        # 処理開始時刻
        now = datetime.now()
        # self.ws.cell(row=2, column=5).value = now
        self.write_data["starttime"] = f'{now}'

        # 保存先ファイル名
        # self.output_file_path = f'{self.OUTPUT_FILE_DIR_PATH}output_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'
        # self.output_file_path = f'output_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'

        # template.xlsxをアップロード
        scrts = st.secrets.connections.gsheets.to_dict()
        # st.write(type(scrts))
        # st.write(scrts)

        # サービスアカウント認証
        gauth = GoogleAuth()
        scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
        gauth.credentials = ServiceAccountCredentials.from_json_keyfile_dict(scrts, scope)
        drive = GoogleDrive(gauth)

        # 保存先のディレクトリID
        ID = "16p7o34MT2WuABNhJfBGeGp5Yb-W-JtLG"

        # 保存先のファイル名
        output_file_name = f'output_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'

        # アップロード処理
        file = drive.CreateFile({"title": output_file_name, "parents": [{"id": ID}]})
        local_file_path = "./output/template.xlsx"
        file.SetContentFile(local_file_path)
        is_convert = True
        file.Upload({'convert': is_convert})
        self._gsheet = file

        # 出力先はgsのリンク先を設定
        self.output_file_path = self._gsheet.get('alternateLink')

        # スプレッドシートの認証
        gc = gspread.authorize(gauth.credentials)
        wb = gc.open_by_key(self._gsheet.get('id'))
        self._ws = wb.get_worksheet(0)

        # シート名を変更
        self._ws.update_title('不動産請求情報一覧')

    def writeCondition(self, user_id, conditions):
        # 収集条件
        condition = f"【実行ユーザ】{user_id}／"
        condition += editCollectionCondition(conditions)
        # self.ws.cell(row=1, column=5).value = condition
        self.write_data["conditions"] = condition
    
    def write(self, data):
        # データを書き込む
        # for col in range(1, len(data)+1):
        #     self.ws.cell(row=self.row_count, column=col+1).value = data[col-1]
        self.write_data["rows"].append(data)
        self.row_count += 1

    def save(self):
        # 処理終了時刻
        # self.ws.cell(row=3, column=5).alue = datetime.now()
        self.write_data["endtime"] = f'{datetime.now()}'

        # 印刷範囲設定
        # self.ws.print_area = f'A1:J{(self.row_count-1)}'
        
        # ToDo:gs更新
        # print(self.write_data)

        # 収集条件
        # self.ws.cell(row=1, column=5, value = self.write_data["conditions"])
        # ws.update_cell(1, 5, self.write_data["conditions"])
        # 収集条件、開始時刻、終了時刻、終了状態
        self._ws.update("E1:F3", [[self.write_data["conditions"]],
                            [self.write_data["starttime"]],
                            [self.write_data["endtime"],
                             self.write_data["state"]]])

        # 開始時刻
        # self.ws.cell(row=2, column=5, value = self.write_data["starttime"])
        # ws.update_cell(2, 5, self.write_data["starttime"])

        # 収集情報
        # for i, data in enumerate(self.write_data["rows"]):
        #     # print(data)
        #     for col in range(1, len(data)+1):
        #         self.ws.cell(row=i+6, column=col+1, value = data[col-1])
        #         # ws.update_cell(i+6, col+1, data[col-1])
        if len(self.write_data["rows"]) > 0:
            self._ws.update(f"B6:J{6+len(self.write_data["rows"])}", self.write_data["rows"])

        # 終了時刻       
        # self.ws.cell(row=3, column=5, value = self.write_data["endtime"])
        # ws.update_cell(3, 5, self.write_data["endtime"])

        # 終了状態
        # self.ws.cell(row=3, column=6, value=self.write_data["state"])
        # ws.update_cell(3, 6, self.write_data["state"])
        
        # Excelファイルを保存
        # self.wb.save(self.output_file_path)
        
    
    def save_with_exception(self):
        # エラー発生を記録
        # self.ws.cell(row=3, column=6).value = '★エラー発生し処理中断しました。エラー対処後に再実行してください。★'
        self.write_data["state"] =  '★エラー発生し処理中断しました。エラー対処後に再実行してください。★'
        # Excelファイルを保存
        self.save()

    def writeFudousan(self, seikyuJiko, shozaiType, shozai):
        # 請求種別／種別／所在及び地番・家屋番号
        self.write([self.data_count, seikyuJiko, shozaiType, shozai])
        self.data_count += 1

    def writeZumen(self, seikyuJiko, shozaiType, shozai, toukiDate, jiken_no, shozai_before, shozai_after):
        # 請求種別／種別／所在及び地番・家屋番号／登録年月日／事件ID／所在及び地番・家屋番号（事件前物件）／所在及び地番・家屋番号（事件後物件）
        self.write([self.data_count, seikyuJiko, shozaiType ,shozai, toukiDate, jiken_no, shozai_before, shozai_after])
        self.data_count += 1

    def writeZemenNasi(self, seikyuJiko, shozaiType, shozai, reasonZumenNashi):
        # 請求種別／種別／所在及び地番・家屋番号／図面なし理由
        self.write([self.data_count, seikyuJiko, shozaiType, shozai, '',  '',  '',  '', reasonZumenNashi])
        self.data_count += 1